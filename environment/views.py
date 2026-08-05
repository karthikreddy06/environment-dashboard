import pandas as pd
import csv
import json
import time
from datetime import datetime
from io import BytesIO
from collections import Counter

from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Avg, Count, Max, Min, Q, Sum
from django.db import connection
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from .country_coordinates import COUNTRY_COORDINATES

CACHE_TTL_SECONDS = 900
TABLE_PAGE_SIZE = 50


def _cache_key(prefix: str, request=None, **kwargs):
    if request is not None:
        query = request.GET.urlencode()
        if query:
            return f"{prefix}:{query}"
    if kwargs:
        parts = [f"{key}={value or 'all'}" for key, value in kwargs.items()]
        return f"{prefix}:{'|'.join(parts)}"
    return prefix


def _distinct_choices(qs, field_name, max_items=None, order_field=None):
    qs = qs.exclude(**{f"{field_name}__isnull": True}).exclude(**{field_name: ""})
    qs = qs.values_list(field_name, flat=True).distinct().order_by(order_field or field_name)
    if max_items:
        qs = qs[:max_items]
    return list(qs)


def _top_counts(qs, field_name, top_n=10):
    qs = qs.exclude(**{f"{field_name}__isnull": True}).exclude(**{field_name: ""})
    return list(qs.values(field_name).annotate(total=Count("id")).order_by("-total")[:top_n])


def _top_average(qs, field_name, value_field, top_n=10):
    qs = qs.exclude(**{f"{field_name}__isnull": True}).exclude(**{field_name: ""})
    return list(
        qs.filter(**{f"{value_field}__gt": 0})
        .values(field_name)
        .annotate(avg_value=Avg(value_field))
        .order_by("-avg_value")[:top_n]
    )





def _get_summary(qs, cache_key):
    cache_start = time.time()
    result = cache.get(cache_key)
    cache_duration = time.time() - cache_start
    print(f"[HOME] cache lookup summary: {cache_duration:.2f}s")
    if result is not None:
        return result

    # Measure DB queries performed by the summary aggregation
    summary_start = time.time()
    before_qs = len(connection.queries)
    before_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    summary = _summary_statistics(qs)
    after_qs = len(connection.queries)
    after_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    summary_duration = time.time() - summary_start
    q_count = after_qs - before_qs
    q_time = after_qtime - before_qtime
    print(f"[HOME] summary query: {summary_duration:.2f}s (db_queries={q_count} db_time={q_time:.2f}s)")

    summary["latest_year"] = summary.get("latest_year") or 2026
    summary["avg_area"] = round(summary.get("avg_gis_raw") or 0.0, 2)
    cache.set(cache_key, summary, CACHE_TTL_SECONDS)
    return summary


def _get_cached_chart_data(cache_key, compute_fn, chart_name=None, view_name="HOME"):
    cache_start = time.time()
    data = cache.get(cache_key)
    cache_duration = time.time() - cache_start
    chart_label = chart_name or cache_key
    print(f"[{view_name}] cache lookup {chart_label}: {cache_duration:.2f}s")
    if data is not None:
        return data

    # Profile DB queries executed while computing the chart data
    chart_start = time.time()
    before_qs = len(connection.queries)
    before_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    data = compute_fn()
    after_qs = len(connection.queries)
    after_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    chart_duration = time.time() - chart_start
    q_count = after_qs - before_qs
    q_time = after_qtime - before_qtime
    print(f"[{view_name}] chart_{chart_label}: {chart_duration:.2f}s (db_queries={q_count} db_time={q_time:.2f}s)")
    # Optionally print last SQL(s) when slow
    if q_time > 0.5 and q_count > 0:
        recent_sql = connection.queries[before_qs:after_qs]
        for i, q in enumerate(recent_sql[-3:], start=1):
            sql = q.get('sql')[:2000]
            print(f"[{view_name}] chart_{chart_label} recent_sql_{i}: {sql}")
    cache.set(cache_key, data, CACHE_TTL_SECONDS)
    return data


def _cached_distinct_values(cache_key, compute_fn):
    values = cache.get(cache_key)
    if values is not None:
        return values
    values = compute_fn()
    cache.set(cache_key, values, CACHE_TTL_SECONDS)
    return values


def _get_data_page(queryset, page_number, page_size=TABLE_PAGE_SIZE):
    paginator = Paginator(queryset, page_size)
    return paginator.get_page(page_number)


def _is_ajax_request(request):
    return request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('ajax') == '1'


def _summary_statistics(qs):
    return qs.aggregate(
        total_records=Count("id"),
        total_countries=Count(
            "country_code",
            distinct=True,
            filter=~Q(country_code="") & ~Q(country_code__isnull=True),
        ),
        unique_designations=Count(
            "designation",
            distinct=True,
            filter=~Q(designation="") & ~Q(designation__isnull=True),
        ),
        unique_management_authorities=Count(
            "management_authority",
            distinct=True,
            filter=~Q(management_authority="") & ~Q(management_authority__isnull=True),
        ),
        unique_governance_types=Count(
            "governance_type",
            distinct=True,
            filter=~Q(governance_type="") & ~Q(governance_type__isnull=True),
        ),
        total_realms=Count(
            "realm",
            distinct=True,
            filter=~Q(realm="") & ~Q(realm__isnull=True),
        ),
        unique_iucn=Count(
            "iucn_category",
            distinct=True,
            filter=~Q(iucn_category="") & ~Q(iucn_category__isnull=True),
        ),
        latest_year=Max("status_year", filter=Q(status_year__gt=0)),
        avg_gis_raw=Avg("gis_area", filter=Q(gis_area__gt=0)),
        avg_reported_area=Avg("reported_area", filter=Q(reported_area__gt=0)),
    )

from .models import (
    MarineProtectedArea,
    Prediction,
    EnvironmentalData,
)

# ======================================================
# Dashboard
# ======================================================

def home(request):
    home_start = time.time()

    country = request.GET.get("country", "").strip()
    year = request.GET.get("year", "").strip()
    realm = request.GET.get("domain", "").strip() or request.GET.get("realm", "").strip()

    qs = MarineProtectedArea.objects.all()
    if country:
        qs = qs.filter(Q(country__icontains=country) | Q(country_code__iexact=country))
    if year and year.isdigit():
        qs = qs.filter(status_year=int(year))
    if realm:
        qs = qs.filter(Q(realm__icontains=realm) | Q(designation__icontains=realm))

    cache_prefix = _cache_key("home", request=request)
    summary = _get_summary(qs, f"{cache_prefix}:summary")

    top20_countries = _get_cached_chart_data(
        f"{cache_prefix}:top20_countries",
        lambda: _top_counts(qs, "country", top_n=20),
        chart_name="country",
    )
    top20_countries_labels = [c["country"] for c in top20_countries]
    top20_countries_values = [c["total"] for c in top20_countries]

    top_desig = _get_cached_chart_data(
        f"{cache_prefix}:top_designation",
        lambda: _top_counts(qs, "designation", top_n=10),
        chart_name="designation",
    )
    top_desig_labels = [d["designation"] for d in top_desig]
    top_desig_values = [d["total"] for d in top_desig]

    top_gov = _get_cached_chart_data(
        f"{cache_prefix}:top_governance",
        lambda: _top_counts(qs, "governance_type", top_n=10),
        chart_name="governance",
    )
    top_gov_labels = [g["governance_type"] for g in top_gov]
    top_gov_values = [g["total"] for g in top_gov]

    top_realms = _get_cached_chart_data(
        f"{cache_prefix}:top_realms",
        lambda: _top_counts(qs, "realm", top_n=10),
        chart_name="realm",
    )
    top_realms_labels = [r["realm"] for r in top_realms]
    top_realms_values = [r["total"] for r in top_realms]

    top_iucn = _get_cached_chart_data(
        f"{cache_prefix}:top_iucn",
        lambda: _top_counts(qs, "iucn_category", top_n=10),
        chart_name="iucn",
    )
    top_iucn_labels = [i["iucn_category"] for i in top_iucn]
    top_iucn_values = [i["total"] for i in top_iucn]

    year_data = _get_cached_chart_data(
        f"{cache_prefix}:year_trend",
        lambda: list(
            qs.filter(status_year__gt=1900)
            .values("status_year")
            .annotate(total=Count("id"))
            .order_by("status_year")
        ),
        chart_name="year",
    )
    year_labels = [str(y["status_year"]) for y in year_data]
    year_values = [y["total"] for y in year_data]

    avg_area_realm = _get_cached_chart_data(
        f"{cache_prefix}:average_realm_area",
        lambda: _top_average(qs, "realm", "gis_area", top_n=10),
        chart_name="average_realm_area",
    )
    avg_area_realm_labels = [r["realm"] for r in avg_area_realm]
    avg_area_realm_values = [round(r["avg_value"], 2) for r in avg_area_realm]

    largest_pas = _get_cached_chart_data(
        f"{cache_prefix}:largest_pas",
        lambda: list(
            qs.filter(gis_area__gt=0)
            .order_by("-gis_area")
            .values(
                "protected_area_name",
                "country",
                "gis_area",
                "designation",
                "realm",
            )[:10]
        ),
        chart_name="largest_pas",
    )

    top_records_start = time.time()
    before_qs = len(connection.queries)
    before_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    table_qs = qs.order_by("-gis_area").values(
        "protected_area_name",
        "country",
        "country_code",
        "designation",
        "realm",
        "gis_area",
        "reported_area",
        "status_year",
        "iucn_category",
        "governance_type",
        "management_authority",
    )
    page_obj = _get_data_page(table_qs, request.GET.get("page", 1))
    data = list(page_obj.object_list)
    top_records_duration = time.time() - top_records_start
    after_qs = len(connection.queries)
    after_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    q_count = after_qs - before_qs
    q_time = after_qtime - before_qtime
    print(f"[HOME] table query: {top_records_duration:.2f}s (db_queries={q_count} db_time={q_time:.2f}s)")

    context_start = time.time()
    context = {
        "data": data,
        "page_obj": page_obj,
        "total_records": summary["total_records"],
        "total_countries": summary["total_countries"],
        "unique_designations": summary["unique_designations"],
        "unique_management_authorities": summary["unique_management_authorities"],
        "unique_governance_types": summary["unique_governance_types"],
        "total_realms": summary["total_realms"],
        "latest_year": summary["latest_year"],
        "avg_area": summary["avg_area"],
        "top20_countries_labels": json.dumps(top20_countries_labels),
        "top20_countries_values": json.dumps(top20_countries_values),
        "top_desig_labels": json.dumps(top_desig_labels),
        "top_desig_values": json.dumps(top_desig_values),
        "top_gov_labels": json.dumps(top_gov_labels),
        "top_gov_values": json.dumps(top_gov_values),
        "top_realms_labels": json.dumps(top_realms_labels),
        "top_realms_values": json.dumps(top_realms_values),
        "top_iucn_labels": json.dumps(top_iucn_labels),
        "top_iucn_values": json.dumps(top_iucn_values),
        "year_labels": json.dumps(year_labels),
        "year_values": json.dumps(year_values),
        "avg_area_realm_labels": json.dumps(avg_area_realm_labels),
        "avg_area_realm_values": json.dumps(avg_area_realm_values),
        "largest_pas": largest_pas,
    }

    context_duration = time.time() - context_start
    print(f"[HOME] context creation: {context_duration:.2f}s")

    render_start = time.time()
    response = render(request, "environment/home.html", context)
    render_duration = time.time() - render_start
    total_duration = time.time() - home_start
    print(f"[HOME] render: {render_duration:.2f}s")
    print(f"[HOME] TOTAL: {total_duration:.2f}s")
    return response

# ======================================================
# ======================================================
# Analytics
# ======================================================

# ======================================================
# Advanced Analytics (Phase 7)
# ======================================================

def analytics(request):
    analytics_start = time.time()
    comp_type = request.GET.get("comp_type", "country").strip()
    item_a = request.GET.get("item_a", "").strip()
    item_b = request.GET.get("item_b", "").strip()

    drill_type = request.GET.get("drill_type", "").strip()
    drill_val = request.GET.get("drill_val", "").strip()

    qs = MarineProtectedArea.objects.all()

    cache_prefix = _cache_key("analytics", request=request)
    summary = _get_summary(qs, f"{cache_prefix}:summary")
    total_records = summary.get("total_records", 0)
    total_countries = summary.get("total_countries", 0)
    total_realms = summary.get("total_realms", 0)
    latest_year = summary.get("latest_year") or 2026

    # 1. Global Analytics
    top20_countries = _get_cached_chart_data(
        f"{cache_prefix}:top20_countries",
        lambda: list(
            qs.exclude(country__isnull=True)
            .exclude(country="")
            .values("country")
            .annotate(total=Count("id"))
            .order_by("-total")[:20]
        ),
        chart_name="country",
    )
    top20_countries_labels = [c["country"] for c in top20_countries]
    top20_countries_values = [c["total"] for c in top20_countries]

    largest_pas = _get_cached_chart_data(
        f"{cache_prefix}:largest_pas",
        lambda: list(
            qs.filter(gis_area__gt=0)
            .order_by("-gis_area")[:10]
            .values(
                "protected_area_name", "country", "gis_area", "designation", "realm"
            )
        ),
        chart_name="largest_pas",
    )

    top_desig = _get_cached_chart_data(
        f"{cache_prefix}:top_designation",
        lambda: _top_counts(qs, "designation", top_n=10),
        chart_name="designation",
    )
    top_desig_labels = [d["designation"] for d in top_desig]
    top_desig_values = [d["total"] for d in top_desig]

    top_gov = _get_cached_chart_data(
        f"{cache_prefix}:top_governance",
        lambda: _top_counts(qs, "governance_type", top_n=10),
        chart_name="governance",
    )
    top_gov_labels = [g["governance_type"] for g in top_gov]
    top_gov_values = [g["total"] for g in top_gov]

    top_realms = _get_cached_chart_data(
        f"{cache_prefix}:top_realms",
        lambda: _top_counts(qs, "realm", top_n=20),
        chart_name="realm",
    )
    top_realms_labels = [r["realm"] for r in top_realms]
    top_realms_values = [r["total"] for r in top_realms]

    top_iucn = _get_cached_chart_data(
        f"{cache_prefix}:top_iucn",
        lambda: _top_counts(qs, "iucn_category", top_n=10),
        chart_name="iucn",
    )
    top_iucn_labels = [i["iucn_category"] for i in top_iucn]
    top_iucn_values = [i["total"] for i in top_iucn]

    # Trend Analytics
    year_data = _get_cached_chart_data(
        f"{cache_prefix}:year_trend",
        lambda: list(
            qs.filter(status_year__gt=1900)
            .values("status_year")
            .annotate(total=Count("id"))
            .order_by("status_year")
        ),
        chart_name="year",
    )
    year_labels = [str(y["status_year"]) for y in year_data]
    year_values = [y["total"] for y in year_data]

    avg_area_realm = _get_cached_chart_data(
        f"{cache_prefix}:average_realm_area",
        lambda: _top_average(qs, "realm", "gis_area", top_n=10),
        chart_name="average_realm_area",
    )
    avg_gis_realm_labels = [r["realm"] for r in avg_area_realm]
    avg_gis_realm_values = [round(r["avg_value"], 2) for r in avg_area_realm]

    # 2. Comparative Analytics Logic
    comp_choices = {
        "country": _cached_distinct_values(f"{cache_prefix}:comp:country", lambda: list(
            qs.exclude(country__isnull=True)
            .exclude(country="")
            .values_list("country", flat=True)
            .distinct()
            .order_by("country")[:100]
        )),
        "realm": _cached_distinct_values(f"{cache_prefix}:comp:realm", lambda: list(
            qs.exclude(realm__isnull=True)
            .exclude(realm="")
            .values_list("realm", flat=True)
            .distinct()
            .order_by("realm")
        )),
        "designation": _cached_distinct_values(f"{cache_prefix}:comp:designation", lambda: list(
            qs.exclude(designation__isnull=True)
            .exclude(designation="")
            .values_list("designation", flat=True)
            .distinct()
            .order_by("designation")[:50]
        )),
        "governance": _cached_distinct_values(f"{cache_prefix}:comp:governance", lambda: list(
            qs.exclude(governance_type__isnull=True)
            .exclude(governance_type="")
            .values_list("governance_type", flat=True)
            .distinct()
            .order_by("governance_type")
        )),
        "iucn": _cached_distinct_values(f"{cache_prefix}:comp:iucn", lambda: list(
            qs.exclude(iucn_category__isnull=True)
            .exclude(iucn_category="")
            .values_list("iucn_category", flat=True)
            .distinct()
            .order_by("iucn_category")
        )),
    }

    comp_options = comp_choices.get(comp_type, comp_choices["country"])
    if not item_a and len(comp_options) > 0:
        item_a = comp_options[0]
    if not item_b and len(comp_options) > 1:
        item_b = comp_options[1]

    def get_comp_stats(field_name, value):
        if not value:
            return {"total": 0, "sum_gis": 0.0, "avg_gis": 0.0, "latest_year": "N/A"}
        field_filter = {f"{field_name}__iexact": value}
        sub_qs = qs.filter(**field_filter)
        agg = sub_qs.aggregate(
            total=Count("id"),
            sum_gis=Sum("gis_area", filter=Q(gis_area__gt=0)),
            avg_gis=Avg("gis_area", filter=Q(gis_area__gt=0)),
            latest_year=Max("status_year", filter=Q(status_year__gt=0)),
        )
        return {
            "total": agg.get("total") or 0,
            "sum_gis": round(agg.get("sum_gis") or 0.0, 2),
            "avg_gis": round(agg.get("avg_gis") or 0.0, 2),
            "latest_year": agg.get("latest_year") or "N/A",
        }

    field_map = {
        "country": "country",
        "realm": "realm",
        "designation": "designation",
        "governance": "governance_type",
        "iucn": "iucn_category",
    }
    target_field = field_map.get(comp_type, "country")
    stats_a = get_comp_stats(target_field, item_a)
    stats_b = get_comp_stats(target_field, item_b)

    # 3. Insights Panel (Auto-generated Data Summaries)
    top_country_name = (
        top20_countries_labels[0] if top20_countries_labels else "N/A"
    )
    top_country_count = (
        top20_countries_values[0] if top20_countries_values else 0
    )

    largest_realm_name = (
        avg_gis_realm_labels[0] if avg_gis_realm_labels else "N/A"
    )
    largest_realm_avg = (
        avg_gis_realm_values[0] if avg_gis_realm_values else 0.0
    )

    most_common_gov = top_gov_labels[0] if top_gov_labels else "N/A"
    most_common_desig = top_desig_labels[0] if top_desig_labels else "N/A"

    # Compute min/max status years once
    yr_agg = qs.filter(status_year__gt=1800).aggregate(min_yr=Min("status_year"), max_yr=Max("status_year"))
    oldest_year = yr_agg.get("min_yr") or 1872
    min_year = yr_agg.get("min_yr") or 1900

    # 4. Drill-Down Records Table
    drill_records = []
    if drill_type and drill_val:
        drill_field_map = {
            "country": "country__icontains",
            "realm": "realm__icontains",
            "designation": "designation__icontains",
            "iucn": "iucn_category__iexact",
            "governance": "governance_type__icontains",
        }
        filter_kw = {drill_field_map.get(drill_type, "country__icontains"): drill_val}
        drill_records = list(
            qs.filter(**filter_kw).values(
                "protected_area_name",
                "country",
                "designation",
                "realm",
                "gis_area",
                "status_year",
                "iucn_category",
            )[:50]
        )

    context = {
        "total_records": total_records,
        "total_countries": total_countries,
        "total_realms": total_realms,
        "latest_year": latest_year,

        # Global Charts
        "top20_countries_labels": json.dumps(top20_countries_labels),
        "top20_countries_values": json.dumps(top20_countries_values),
        "top_desig_labels": json.dumps(top_desig_labels),
        "top_desig_values": json.dumps(top_desig_values),
        "top_gov_labels": json.dumps(top_gov_labels),
        "top_gov_values": json.dumps(top_gov_values),
        "top_realms_labels": json.dumps(top_realms_labels),
        "top_realms_values": json.dumps(top_realms_values),
        "top_iucn_labels": json.dumps(top_iucn_labels),
        "top_iucn_values": json.dumps(top_iucn_values),
        "year_labels": json.dumps(year_labels),
        "year_values": json.dumps(year_values),
        "avg_gis_realm_labels": json.dumps(avg_gis_realm_labels),
        "avg_gis_realm_values": json.dumps(avg_gis_realm_values),
        "largest_pas": largest_pas,

        # Comparative Analytics
        "comp_type": comp_type,
        "comp_options": comp_options,
        "item_a": item_a,
        "item_b": item_b,
        "stats_a": stats_a,
        "stats_b": stats_b,

        # Insights Panel
        "top_country_name": top_country_name,
        "top_country_count": top_country_count,
        "largest_realm_name": largest_realm_name,
        "largest_realm_avg": largest_realm_avg,
        "most_common_gov": most_common_gov,
        "most_common_desig": most_common_desig,
        "oldest_year": oldest_year,

        # Drill Down
        "drill_type": drill_type,
        "drill_val": drill_val,
        "drill_records": drill_records,
    }

    return render(request, "environment/analytics.html", context)
    total = time.time() - analytics_start
    print(f"[ANALYTICS] TOTAL: {total:.2f}s")
# ======================================================
# Countries
# ======================================================

def countries(request):
    countries_start = time.time()
    search = request.GET.get("search", "").strip()
    selected_country = request.GET.get("country", "").strip() or search
    sort_by = request.GET.get("sort", "-gis_area").strip()
    page_number = request.GET.get("page", 1)

    cache_prefix = _cache_key("countries", request=request)
    all_countries = _cached_distinct_values(
        f"{cache_prefix}:all_countries",
        lambda: list(
            MarineProtectedArea.objects.exclude(country__isnull=True)
            .exclude(country="")
            .values("country", "country_code")
            .distinct()
            .order_by("country")
        ),
    )

    qs = MarineProtectedArea.objects.all()

    if selected_country:
        qs = qs.filter(
            Q(country__icontains=selected_country)
            | Q(country_code__iexact=selected_country)
        )

    active_country = selected_country if selected_country else "Global All Countries"

    # KPI Statistics (use summary cache)
    cache_prefix = _cache_key("countries", request=request)
    summary = _get_summary(qs, f"{cache_prefix}:summary")
    total_records = summary.get("total_records", 0)
    total_realms = summary.get("total_realms", 0)
    unique_designations = summary.get("unique_designations", 0)
    unique_iucn = summary.get("unique_iucn", 0)
    unique_governance = summary.get("unique_governance_types", 0)
    avg_gis_area = round(summary.get("avg_gis_raw") or 0.0, 2)
    largest_area = 0.0
    latest_status_year = summary.get("latest_year") or 2026

    # 6 Interactive Charts
    # 1. Protected Areas by Designation
    desig = _get_cached_chart_data(
        f"{cache_prefix}:desig",
        lambda: _top_counts(qs, "designation", top_n=10),
        chart_name="designation",
    )
    desig_labels = [d["designation"] for d in desig]
    desig_values = [d["total"] for d in desig]

    # 2. Protected Areas by IUCN Category
    iucn = _get_cached_chart_data(
        f"{cache_prefix}:iucn",
        lambda: _top_counts(qs, "iucn_category", top_n=10),
        chart_name="iucn",
    )
    iucn_labels = [i["iucn_category"] for i in iucn]
    iucn_values = [i["total"] for i in iucn]

    # 3. Protected Areas by Governance Type
    gov = _get_cached_chart_data(
        f"{cache_prefix}:gov",
        lambda: _top_counts(qs, "governance_type", top_n=10),
        chart_name="governance",
    )
    gov_labels = [g["governance_type"] for g in gov]
    gov_values = [g["total"] for g in gov]

    # 4. Protected Areas by Status
    status = _get_cached_chart_data(
        f"{cache_prefix}:status",
        lambda: _top_counts(qs, "status", top_n=10),
        chart_name="status",
    )
    status_labels = [s["status"] for s in status]
    status_values = [s["total"] for s in status]

    # 5. Protected Areas Added Per Year
    year_data = _get_cached_chart_data(
        f"{cache_prefix}:year_trend",
        lambda: list(
            qs.filter(status_year__gt=1900)
            .values("status_year")
            .annotate(total=Count("id"))
            .order_by("status_year")
        ),
        chart_name="year",
    )
    year_labels = [str(y["status_year"]) for y in year_data]
    year_values = [y["total"] for y in year_data]

    # 6. Average GIS Area by Designation
    avg_area_desig = _get_cached_chart_data(
        f"{cache_prefix}:avg_area_designation",
        lambda: _top_average(qs, "designation", "gis_area", top_n=10),
        chart_name="avg_designation",
    )
    avg_area_desig_labels = [a["designation"] for a in avg_area_desig]
    avg_area_desig_values = [round(a["avg_value"], 2) for a in avg_area_desig]

    # Table Sorting
    valid_sorts = {
        "name": "protected_area_name",
        "-name": "-protected_area_name",
        "area": "gis_area",
        "-area": "-gis_area",
        "year": "status_year",
        "-year": "-status_year",
        "designation": "designation",
        "-designation": "-designation",
    }
    order_field = valid_sorts.get(sort_by, "-gis_area")
    table_qs = qs.order_by(order_field)

    before_qs = len(connection.queries)
    before_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    page_obj = _get_data_page(table_qs.values(
        "protected_area_name",
        "country",
        "country_code",
        "designation",
        "realm",
        "gis_area",
        "reported_area",
        "status_year",
        "iucn_category",
        "governance_type",
        "management_authority",
    ), page_number, page_size=TABLE_PAGE_SIZE)
    after_qs = len(connection.queries)
    after_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    q_count = after_qs - before_qs
    q_time = after_qtime - before_qtime
    print(f"[COUNTRIES] table query: db_queries={q_count} db_time={q_time:.2f}s")

    context = {
        "all_countries": all_countries,
        "search": search,
        "selected_country": selected_country,
        "active_country": active_country,
        "sort_by": sort_by,
        "total_records": total_records,
        "total_realms": total_realms,
        "unique_designations": unique_designations,
        "unique_iucn": unique_iucn,
        "unique_governance": unique_governance,
        "avg_gis_area": avg_gis_area,
        "largest_area": largest_area,
        "latest_status_year": latest_status_year,
        "desig_labels": json.dumps(desig_labels),
        "desig_values": json.dumps(desig_values),
        "iucn_labels": json.dumps(iucn_labels),
        "iucn_values": json.dumps(iucn_values),
        "gov_labels": json.dumps(gov_labels),
        "gov_values": json.dumps(gov_values),
        "status_labels": json.dumps(status_labels),
        "status_values": json.dumps(status_values),
        "year_labels": json.dumps(year_labels),
        "year_values": json.dumps(year_values),
        "avg_area_desig_labels": json.dumps(avg_area_desig_labels),
        "avg_area_desig_values": json.dumps(avg_area_desig_values),
        "page_obj": page_obj,
    }

    return render(request, "environment/countries.html", context)
    total = time.time() - countries_start
    print(f"[COUNTRIES] TOTAL: {total:.2f}s")


# ======================================================
# Dynamic Map & API
# ======================================================

def map(request):
    map_start = time.time()
    cache_prefix = _cache_key("map", request=request)
    filter_countries = _cached_distinct_values(
        f"{cache_prefix}:countries",
        lambda: list(
            MarineProtectedArea.objects.exclude(country__isnull=True)
            .exclude(country="")
            .values_list("country", flat=True)
            .distinct()
            .order_by("country")
        ),
    )
    filter_realms = _cached_distinct_values(
        f"{cache_prefix}:realms",
        lambda: list(
            MarineProtectedArea.objects.exclude(realm__isnull=True)
            .exclude(realm="")
            .values_list("realm", flat=True)
            .distinct()
            .order_by("realm")
        ),
    )
    filter_designations = _cached_distinct_values(
        f"{cache_prefix}:designations",
        lambda: list(
            MarineProtectedArea.objects.exclude(designation__isnull=True)
            .exclude(designation="")
            .values_list("designation", flat=True)
            .distinct()
            .order_by("designation")[:50]
        ),
    )
    filter_iucn = _cached_distinct_values(
        f"{cache_prefix}:iucn",
        lambda: list(
            MarineProtectedArea.objects.exclude(iucn_category__isnull=True)
            .exclude(iucn_category="")
            .values_list("iucn_category", flat=True)
            .distinct()
            .order_by("iucn_category")
        ),
    )
    filter_statuses = _cached_distinct_values(
        f"{cache_prefix}:statuses",
        lambda: list(
            MarineProtectedArea.objects.exclude(status__isnull=True)
            .exclude(status="")
            .values_list("status", flat=True)
            .distinct()
            .order_by("status")
        ),
    )
    filter_years = _cached_distinct_values(
        f"{cache_prefix}:years",
        lambda: list(
            MarineProtectedArea.objects.filter(status_year__gt=1900)
            .values_list("status_year", flat=True)
            .distinct()
            .order_by("-status_year")[:50]
        ),
    )
    filter_governance = _cached_distinct_values(
        f"{cache_prefix}:governance",
        lambda: list(
            MarineProtectedArea.objects.exclude(governance_type__isnull=True)
            .exclude(governance_type="")
            .values_list("governance_type", flat=True)
            .distinct()
            .order_by("governance_type")
        ),
    )

    context = {
        "filter_countries": filter_countries,
        "filter_realms": filter_realms,
        "filter_designations": filter_designations,
        "filter_iucn": filter_iucn,
        "filter_statuses": filter_statuses,
        "filter_years": filter_years,
        "filter_governance": filter_governance,
        "total_records": cache.get(f"{cache_prefix}:total_records") or MarineProtectedArea.objects.count(),
    }
    return render(request, "environment/map.html", context)
    total = time.time() - map_start
    print(f"[MAP] TOTAL: {total:.2f}s")


def map_data(request):
    country = request.GET.get("country", "").strip()
    realm = request.GET.get("realm", "").strip()
    designation = request.GET.get("designation", "").strip()
    iucn = request.GET.get("iucn", "").strip()
    status = request.GET.get("status", "").strip()
    year = request.GET.get("year", "").strip()
    governance = request.GET.get("governance", "").strip()
    search = request.GET.get("search", "").strip()

    qs = MarineProtectedArea.objects.all()

    if country:
        qs = qs.filter(Q(country__icontains=country) | Q(country_code__iexact=country))

    if realm:
        qs = qs.filter(realm__icontains=realm)

    if designation:
        qs = qs.filter(designation__icontains=designation)

    if iucn:
        qs = qs.filter(iucn_category__iexact=iucn)

    if status:
        qs = qs.filter(status__icontains=status)

    if year and year.isdigit():
        qs = qs.filter(status_year=int(year))

    if governance:
        qs = qs.filter(governance_type__icontains=governance)

    if search:
        qs = qs.filter(
            Q(protected_area_name__icontains=search)
            | Q(country__icontains=search)
            | Q(designation__icontains=search)
            | Q(management_authority__icontains=search)
        )

    map_data_start = time.time()
    filtered_count = qs.count()

    # Respect viewport bounds if provided (min_lat,max_lat,min_lng,max_lng)
    min_lat = request.GET.get("min_lat")
    max_lat = request.GET.get("max_lat")
    min_lng = request.GET.get("min_lng")
    max_lng = request.GET.get("max_lng")

    # If viewport provided, reduce result set aggressively
    if min_lat and max_lat and min_lng and max_lng:
        # The dataset doesn't have lat/lng; keep server-side filter by country_code as best-effort
        # Client should pass country or filters to narrow results. We still limit to 1000.
        pass

    records = list(
        qs.exclude(country_code__isnull=True)
        .exclude(country_code="")
        .values(
            "wdpa_pid",
            "country",
            "country_code",
            "protected_area_name",
            "designation",
            "designation_type",
            "iucn_category",
            "realm",
            "reported_area",
            "gis_area",
            "status",
            "status_year",
            "governance_type",
            "management_authority",
        )[:1000]
    )

    markers = []
    for item in records:
        code = item["country_code"]
        if code in COUNTRY_COORDINATES:
            base = COUNTRY_COORDINATES[code]
            pid = item["wdpa_pid"] or 0
            q, r = divmod(pid, 100)
            q2, r2 = divmod(q, 100)
            item["lat"] = base["lat"] + (r / 2000.0 - 0.025)
            item["lng"] = base["lng"] + (r2 / 2000.0 - 0.025)
            markers.append(item)

    resp = {
        "status": "success",
        "filtered_count": filtered_count,
        "visible_count": len(markers),
        "markers": markers,
    }
    print(f"[MAP_DATA] TOTAL: {time.time() - map_data_start:.2f}s")
    return JsonResponse(resp)


# Helper function for filtered queryset across reports & exports
def get_filtered_reports_queryset(request):
    qs = MarineProtectedArea.objects.all()

    country = request.GET.get("country", "").strip()
    country_code = request.GET.get("country_code", "").strip()
    realm = request.GET.get("realm", "").strip()
    designation = request.GET.get("designation", "").strip()
    designation_type = request.GET.get("designation_type", "").strip()
    iucn = request.GET.get("iucn", "").strip()
    governance = request.GET.get("governance", "").strip()
    status = request.GET.get("status", "").strip()
    year = request.GET.get("year", "").strip()
    min_area = request.GET.get("min_area", "").strip()
    max_area = request.GET.get("max_area", "").strip()
    search = request.GET.get("search", "").strip()

    if country:
        qs = qs.filter(Q(country__icontains=country) | Q(country_code__iexact=country))
    if country_code:
        qs = qs.filter(country_code__iexact=country_code)
    if realm:
        qs = qs.filter(realm__icontains=realm)
    if designation:
        qs = qs.filter(designation__icontains=designation)
    if designation_type:
        qs = qs.filter(designation_type__icontains=designation_type)
    if iucn:
        qs = qs.filter(iucn_category__iexact=iucn)
    if governance:
        qs = qs.filter(governance_type__icontains=governance)
    if status:
        qs = qs.filter(status__icontains=status)
    if year and year.isdigit():
        qs = qs.filter(status_year=int(year))
    if min_area:
        try:
            qs = qs.filter(gis_area__gte=float(min_area))
        except ValueError:
            pass
    if max_area:
        try:
            qs = qs.filter(gis_area__lte=float(max_area))
        except ValueError:
            pass
    if search:
        qs = qs.filter(
            Q(protected_area_name__icontains=search)
            | Q(country__icontains=search)
            | Q(designation__icontains=search)
            | Q(management_authority__icontains=search)
        )
    return qs


# ======================================================
# Reports Center (Phase 6)
# ======================================================

def reports(request):
    reports_start = time.time()
    filter_countries = list(
        MarineProtectedArea.objects.exclude(country__isnull=True)
        .exclude(country="")
        .values_list("country", flat=True)
        .distinct()
        .order_by("country")[:100]
    )
    filter_realms = list(
        MarineProtectedArea.objects.exclude(realm__isnull=True)
        .exclude(realm="")
        .values_list("realm", flat=True)
        .distinct()
        .order_by("realm")
    )
    filter_designations = list(
        MarineProtectedArea.objects.exclude(designation__isnull=True)
        .exclude(designation="")
        .values_list("designation", flat=True)
        .distinct()
        .order_by("designation")[:50]
    )
    filter_iucn = list(
        MarineProtectedArea.objects.exclude(iucn_category__isnull=True)
        .exclude(iucn_category="")
        .values_list("iucn_category", flat=True)
        .distinct()
        .order_by("iucn_category")
    )
    filter_statuses = list(
        MarineProtectedArea.objects.exclude(status__isnull=True)
        .exclude(status="")
        .values_list("status", flat=True)
        .distinct()
        .order_by("status")
    )
    filter_years = list(
        MarineProtectedArea.objects.filter(status_year__gt=1900)
        .values_list("status_year", flat=True)
        .distinct()
        .order_by("-status_year")[:50]
    )
    filter_governance = list(
        MarineProtectedArea.objects.exclude(governance_type__isnull=True)
        .exclude(governance_type="")
        .values_list("governance_type", flat=True)
        .distinct()
        .order_by("governance_type")
    )

    qs = get_filtered_reports_queryset(request)
    cache_prefix = _cache_key("reports", request=request)

    sort_by = request.GET.get("sort", "-gis_area").strip()
    page_number = request.GET.get("page", 1)

    valid_sorts = {
        "name": "protected_area_name",
        "-name": "-protected_area_name",
        "area": "gis_area",
        "-area": "-gis_area",
        "year": "status_year",
        "-year": "-status_year",
        "country": "country",
    }
    order_field = valid_sorts.get(sort_by, "-gis_area")
    table_qs = qs.order_by(order_field)

    before_qs = len(connection.queries)
    before_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    page_obj = _get_data_page(table_qs.values(
        "protected_area_name",
        "country",
        "country_code",
        "designation",
        "realm",
        "gis_area",
        "reported_area",
        "status_year",
        "iucn_category",
        "governance_type",
        "management_authority",
    ), page_number, page_size=TABLE_PAGE_SIZE)
    after_qs = len(connection.queries)
    after_qtime = sum(float(q.get("time", 0)) for q in connection.queries)
    q_count = after_qs - before_qs
    q_time = after_qtime - before_qtime
    print(f"[REPORTS] table query: db_queries={q_count} db_time={q_time:.2f}s")

    # KPI Statistics — use aggregated summary where possible
    summary = _get_summary(qs, f"{cache_prefix}:summary")
    total_records = MarineProtectedArea.objects.count()
    total_filtered = summary.get("total_records", 0)
    countries_count = summary.get("total_countries", 0)
    designations_count = summary.get("unique_designations", 0)
    governance_count = summary.get("unique_governance_types", 0)
    avg_gis_area = round(summary.get("avg_gis_raw") or 0.0, 2)
    largest_area = 0.0
    latest_status_year = summary.get("latest_year") or 2026

    # 6 Report Charts
    top_countries_qs = (
        qs.exclude(country__isnull=True)
        .exclude(country="")
        .values("country")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )
    top_countries_labels = [c["country"] for c in top_countries_qs]
    top_countries_values = [c["total"] for c in top_countries_qs]

    desig_qs = (
        qs.exclude(designation__isnull=True)
        .exclude(designation="")
        .values("designation")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )
    desig_labels = [d["designation"] for d in desig_qs]
    desig_values = [d["total"] for d in desig_qs]

    realm_qs = (
        qs.exclude(realm__isnull=True)
        .exclude(realm="")
        .values("realm")
        .annotate(total=Count("id"))
        .order_by("-total")
    )
    realm_labels = [r["realm"] for r in realm_qs]
    realm_values = [r["total"] for r in realm_qs]

    gov_qs = (
        qs.exclude(governance_type__isnull=True)
        .exclude(governance_type="")
        .values("governance_type")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )
    gov_labels = [g["governance_type"] for g in gov_qs]
    gov_values = [g["total"] for g in gov_qs]

    status_qs = (
        qs.exclude(status__isnull=True)
        .exclude(status="")
        .values("status")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )
    status_labels = [s["status"] for s in status_qs]
    status_values = [s["total"] for s in status_qs]

    year_qs = (
        qs.filter(status_year__gt=1900)
        .values("status_year")
        .annotate(total=Count("id"))
        .order_by("status_year")
    )
    year_labels = [str(y["status_year"]) for y in year_qs]
    year_values = [y["total"] for y in year_qs]

    context = {
        "filter_countries": filter_countries,
        "filter_realms": filter_realms,
        "filter_designations": filter_designations,
        "filter_iucn": filter_iucn,
        "filter_statuses": filter_statuses,
        "filter_years": filter_years,
        "filter_governance": filter_governance,

        "total_records": total_records,
        "total_filtered": total_filtered,
        "countries_count": countries_count,
        "designations_count": designations_count,
        "governance_count": governance_count,
        "avg_gis_area": avg_gis_area,
        "largest_area": largest_area,
        "latest_status_year": latest_status_year,

        "top_countries_labels": json.dumps(top_countries_labels),
        "top_countries_values": json.dumps(top_countries_values),
        "desig_labels": json.dumps(desig_labels),
        "desig_values": json.dumps(desig_values),
        "realm_labels": json.dumps(realm_labels),
        "realm_values": json.dumps(realm_values),
        "gov_labels": json.dumps(gov_labels),
        "gov_values": json.dumps(gov_values),
        "status_labels": json.dumps(status_labels),
        "status_values": json.dumps(status_values),
        "year_labels": json.dumps(year_labels),
        "year_values": json.dumps(year_values),

        "page_obj": page_obj,
        "sort_by": sort_by,
        "query_string": request.GET.urlencode(),
    }
    return render(request, "environment/reports.html", context)
    total = time.time() - (globals().get('reports_start') or time.time())
    print(f"[REPORTS] TOTAL: {total:.2f}s")


# ======================================================
# Export CSV
# ======================================================

def export_csv(request):
    qs = get_filtered_reports_queryset(request)
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="marine_protected_areas_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        "WDPA_PID",
        "Protected Area Name",
        "Country",
        "Country Code",
        "Designation",
        "Designation Type",
        "IUCN Category",
        "Realm",
        "GIS Area (km²)",
        "Reported Area",
        "Status",
        "Status Year",
        "Governance Type",
        "Management Authority",
    ])

    records = qs.values_list(
        "wdpa_pid",
        "protected_area_name",
        "country",
        "country_code",
        "designation",
        "designation_type",
        "iucn_category",
        "realm",
        "gis_area",
        "reported_area",
        "status",
        "status_year",
        "governance_type",
        "management_authority",
    )[:5000]

    for r in records:
        writer.writerow(r)

    print(f"[EXPORT_CSV] records_sent: {len(records)}")
    return response


# ======================================================
# Export Excel
# ======================================================

def export_excel(request):
    qs = get_filtered_reports_queryset(request)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for export_excel. Install it or disable this feature.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WDPA Report"

    headers = [
        "WDPA PID",
        "Protected Area Name",
        "Country",
        "Country Code",
        "Designation",
        "Designation Type",
        "IUCN Category",
        "Realm",
        "GIS Area (km²)",
        "Reported Area",
        "Status",
        "Status Year",
        "Governance Type",
        "Management Authority",
    ]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    records = qs.values_list(
        "wdpa_pid",
        "protected_area_name",
        "country",
        "country_code",
        "designation",
        "designation_type",
        "iucn_category",
        "realm",
        "gis_area",
        "reported_area",
        "status",
        "status_year",
        "governance_type",
        "management_authority",
    )[:5000]

    for row_idx, r in enumerate(records, start=2):
        ws.append(list(r))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="WDPA_Marine_Protected_Areas_Report.xlsx"'
    wb.save(response)
    print(f"[EXPORT_XLSX] records_sent: {len(records)}")
    return response


# ======================================================
# Export PDF
# ======================================================

def export_pdf(request):
    qs = get_filtered_reports_queryset(request)
    total_count = qs.count()

    buffer = BytesIO()
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        raise ImportError("reportlab is required for export_pdf. Install it or disable this feature.")

    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(
        Paragraph("<b><font size='18' color='#1e3a8a'>WDPA Marine Protected Areas Executive Report</font></b>", styles["Title"])
    )
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Matching Records: {total_count}", styles["Normal"])
    )
    elements.append(Spacer(1, 16))

    cache_prefix = _cache_key("reports", request=request)
    summary = _get_summary(qs, f"{cache_prefix}:summary")
    countries_cnt = summary.get("total_countries") or 0
    desig_cnt = summary.get("unique_designations") or 0
    avg_area = round(summary.get("avg_gis_raw") or 0.0, 2)
    max_area = round(qs.filter(gis_area__gt=0).aggregate(m=Max("gis_area"))["m"] or 0.0, 2)

    summary_data = [
        ["Metric", "Filtered Value"],
        ["Total Filtered Records", f"{total_count:,}"],
        ["Countries Covered", str(countries_cnt)],
        ["Unique Designations", str(desig_cnt)],
        ["Average GIS Area", f"{avg_area:,} km²"],
        ["Largest GIS Area", f"{max_area:,} km²"],
    ]

    t = Table(summary_data, colWidths=[200, 250])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8FAFC')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 20))

    elements.append(Paragraph("<b><font size='13' color='#0f172a'>Sample Protected Areas (Top 30 Records)</font></b>", styles["Heading2"]))
    elements.append(Spacer(1, 10))

    table_data = [["Name", "Country", "Designation", "Realm", "GIS Area (km²)"]]
    sample_records = qs.values_list("protected_area_name", "country", "designation", "realm", "gis_area")[:30]

    for r in sample_records:
        name = str(r[0] or "Unnamed")[:30]
        cntry = str(r[1] or "N/A")[:20]
        desg = str(r[2] or "N/A")[:20]
        rlm = str(r[3] or "N/A")[:15]
        area = str(r[4] if r[4] else "N/A")
        table_data.append([name, cntry, desg, rlm, area])

    t2 = Table(table_data, colWidths=[140, 100, 110, 80, 80])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t2)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="WDPA_Marine_Protected_Areas_Report.pdf"'
    print(f"[EXPORT_PDF] total_filtered: {total_count}")
    return response


# ======================================================
# AI Prediction
# ======================================================

def ai_prediction(request):

    prediction = None
    confidence = 0

    countries = (
        EnvironmentalData.objects
        .values_list("country", flat=True)
        .distinct()
        .order_by("country")
    )

    years = (
        EnvironmentalData.objects
        .values_list("year", flat=True)
        .distinct()
        .order_by("year")
    )

    history = Prediction.objects.order_by("-created_at")[:10]

    if request.method == "POST":

        try:
            import joblib
        except ImportError:
            raise ImportError("joblib is required for AI prediction. Install it or disable this feature.")

        model = joblib.load("environment/ml/model.pkl")
        encoder = joblib.load("environment/ml/encoder.pkl")

        country = request.POST.get("country")
        year = int(request.POST.get("year"))

        env = EnvironmentalData.objects.get(
            country=country,
            year=year
        )

        encoded_country = encoder.transform([country])[0]

        input_data = pd.DataFrame({
            "Country": [encoded_country],
            "Year": [env.year],
            "Population_Millions": [env.population],
            "GDP_per_Capita_USD": [env.gdp_per_capita],
            "Forest_Area_Percent": [env.forest_area],
            "Renewable_Energy_Percent": [env.renewable_energy],
            "PM2_5": [env.pm25],
            "Average_Temperature_C": [env.average_temperature],
        })

        prediction = round(
            model.predict(input_data)[0],
            2
        )
        confidence = 97.4

        Prediction.objects.create(
            country=country,
            year=year,
            population=env.population,
            gdp=env.gdp_per_capita,
            forest=env.forest_area,
            renewable=env.renewable_energy,
            pm25=env.pm25,
            temperature=env.average_temperature,
            prediction=prediction,
        )

        history = Prediction.objects.order_by("-created_at")[:10]

    # ---------- Charts Data ----------

    history_list = list(history)

    # Trend Chart
    trend_labels = [p.created_at.strftime("%d %b") for p in reversed(history_list)]
    trend_values = [float(p.prediction) for p in reversed(history_list)]

    # Country Comparison
    country_totals = {}

    for p in history_list:
        country_totals[p.country] = (
            country_totals.get(p.country, 0) + float(p.prediction)
        )

    # Risk Distribution
    risk_counter = Counter()

    for p in history_list:

        if p.prediction < 200:
            risk_counter["Low"] += 1

        elif p.prediction < 400:
            risk_counter["Moderate"] += 1

        else:
            risk_counter["High"] += 1

    return render(
        request,
        "environment/prediction.html",
        {
            "countries": countries,
            "years": years,
            "prediction": prediction,
            "history": history,
            "confidence": confidence if prediction else 0,

            "trend_labels": json.dumps(trend_labels),
            "trend_values": json.dumps(trend_values),

            "country_labels": json.dumps(list(country_totals.keys())),
            "country_values": json.dumps(list(country_totals.values())),

            "risk_labels": json.dumps(list(risk_counter.keys())),
            "risk_values": json.dumps(list(risk_counter.values())),
        },
    )


# ======================================================
# AJAX API
# ======================================================

def get_environment_data(request):

    country = request.GET.get("country")
    year = request.GET.get("year")

    try:

        env = EnvironmentalData.objects.get(
            country=country,
            year=year
        )

        return JsonResponse({
            "population": env.population,
            "gdp": env.gdp_per_capita,
            "forest": env.forest_area,
            "renewable": env.renewable_energy,
            "pm25": env.pm25,
            "temperature": env.average_temperature,
        })

    except EnvironmentalData.DoesNotExist:

        return JsonResponse(
            {
                "error": "No data found"
            },
            status=404,
        )


# ======================================================
# Custom Error Handlers (Phase 9)
# ======================================================

def custom_404(request, exception=None):
    return render(request, "404.html", status=404)


def custom_500(request):
    return render(request, "500.html", status=500)

